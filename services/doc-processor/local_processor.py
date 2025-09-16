#!/usr/bin/env python3
"""
Local Document Processor - Fallback for Docker issues
Minimal dependencies, runs directly on host
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
import os
import PyPDF2
import docx
import openpyxl
from datetime import datetime
import hashlib
import json

app = Flask(__name__)
CORS(app)

# Configuration
UPLOAD_FOLDER = './data/uploads'
PROCESSED_FOLDER = './data/processed'
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

def extract_text_from_pdf(file_path):
    """Extract text from PDF file"""
    text = []
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                text.append(page.extract_text())
    except Exception as e:
        print(f"Error extracting PDF: {e}")
        return ""
    return "\n".join(text)

def extract_text_from_docx(file_path):
    """Extract text from DOCX file"""
    try:
        doc = docx.Document(file_path)
        text = []
        for paragraph in doc.paragraphs:
            text.append(paragraph.text)
        return "\n".join(text)
    except Exception as e:
        print(f"Error extracting DOCX: {e}")
        return ""

def extract_text_from_xlsx(file_path):
    """Extract text from XLSX file"""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        text = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text.append(f"Sheet: {sheet_name}\n")
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell else "" for cell in row])
                if row_text.strip():
                    text.append(row_text)
        return "\n".join(text)
    except Exception as e:
        print(f"Error extracting XLSX: {e}")
        return ""

@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({
        "status": "healthy",
        "service": "Local Document Processor",
        "mode": "fallback",
        "features": ["pdf", "docx", "xlsx", "txt"]
    })

@app.route('/process', methods=['POST'])
def process_document():
    """Process uploaded document"""
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    # Check file size
    file.seek(0, os.SEEK_END)
    file_size = file.tell()
    file.seek(0)

    if file_size > MAX_FILE_SIZE:
        return jsonify({"error": f"File too large. Max size: {MAX_FILE_SIZE/1024/1024}MB"}), 400

    # Save file
    filename = file.filename
    file_ext = os.path.splitext(filename)[1].lower()

    # Generate unique filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_hash = hashlib.md5(file.read()).hexdigest()[:8]
    file.seek(0)

    saved_filename = f"{timestamp}_{file_hash}{file_ext}"
    file_path = os.path.join(UPLOAD_FOLDER, saved_filename)
    file.save(file_path)

    # Extract text based on file type
    extracted_text = ""
    metadata = {
        "filename": filename,
        "size": file_size,
        "type": file_ext[1:] if file_ext else "unknown",
        "processed_at": datetime.now().isoformat()
    }

    try:
        if file_ext == '.pdf':
            extracted_text = extract_text_from_pdf(file_path)
            metadata["pages"] = len(PyPDF2.PdfReader(open(file_path, 'rb')).pages)
        elif file_ext == '.docx':
            extracted_text = extract_text_from_docx(file_path)
        elif file_ext == '.xlsx':
            extracted_text = extract_text_from_xlsx(file_path)
        elif file_ext in ['.txt', '.md']:
            with open(file_path, 'r', encoding='utf-8') as f:
                extracted_text = f.read()
        else:
            return jsonify({"error": f"Unsupported file type: {file_ext}"}), 400

        # Simple text chunking (1000 chars per chunk with 100 char overlap)
        chunks = []
        chunk_size = 1000
        overlap = 100

        if len(extracted_text) <= chunk_size:
            chunks = [extracted_text]
        else:
            for i in range(0, len(extracted_text), chunk_size - overlap):
                chunk = extracted_text[i:i + chunk_size]
                chunks.append(chunk)

        # Save processed result
        result = {
            "text": extracted_text,
            "chunks": chunks,
            "metadata": metadata
        }

        output_file = os.path.join(PROCESSED_FOLDER, f"{saved_filename}.json")
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

        return jsonify({
            "success": True,
            "text_length": len(extracted_text),
            "chunks_count": len(chunks),
            "metadata": metadata,
            "processed_file": output_file
        })

    except Exception as e:
        return jsonify({"error": f"Processing failed: {str(e)}"}), 500
    finally:
        # Clean up uploaded file
        if os.path.exists(file_path):
            os.remove(file_path)

if __name__ == '__main__':
    print("Starting Local Document Processor on port 5003...")
    print("This is a fallback processor with minimal dependencies")
    app.run(host='0.0.0.0', port=5003, debug=False)